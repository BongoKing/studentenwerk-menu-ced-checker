import json
from pathlib import Path
from openpyxl import load_workbook


def load_settings(path: Path) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def load_allergen_config(path: Path, mode: str) -> dict[str, dict]:
    """Load allergen/additive ratings from xlsx.

    Returns dict: code -> {"beschreibung": str, "bewertung": str}
    mode: "crohn" or "colitis"
    """
    col_name = "Bewertung_Crohn" if mode == "crohn" else "Bewertung_Colitis"

    wb = load_workbook(path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [str(h).strip() if h else "" for h in rows[0]]
    try:
        code_idx = headers.index("Code")
        desc_idx = headers.index("Beschreibung")
        rating_idx = headers.index(col_name)
    except ValueError as e:
        print(f"[!] Spalte nicht gefunden in {path}: {e}")
        return {}

    result = {}
    for row in rows[1:]:
        code = str(row[code_idx]).strip() if row[code_idx] else ""
        desc = str(row[desc_idx]).strip() if row[desc_idx] else ""
        rating = str(row[rating_idx]).strip().lower() if row[rating_idx] else "akzeptabel"
        if code:
            result[code] = {"beschreibung": desc, "bewertung": rating}

    wb.close()
    return result


def load_food_config(path: Path, mode: str) -> dict[str, dict]:
    """Load food ratings from xlsx.

    Returns dict: nahrungsmittel_name (lowered) -> {"bewertung": str}
    mode: "crohn" or "colitis"
    """
    col_name = "Bewertung_Crohn" if mode == "crohn" else "Bewertung_Colitis"

    wb = load_workbook(path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [str(h).strip() if h else "" for h in rows[0]]
    try:
        name_idx = headers.index("Nahrungsmittel")
        rating_idx = headers.index(col_name)
    except ValueError as e:
        print(f"[!] Spalte nicht gefunden in {path}: {e}")
        return {}

    result = {}
    for row in rows[1:]:
        name = str(row[name_idx]).strip() if row[name_idx] else ""
        rating = str(row[rating_idx]).strip().lower() if row[rating_idx] else "akzeptabel"
        if name:
            result[name.lower()] = {"name": name, "bewertung": rating}

    wb.close()
    return result


def load_allergen_config_both(path: Path) -> dict[str, dict]:
    """Load allergen config with both Crohn and Colitis ratings in one pass.

    Returns dict: code -> {"beschreibung": str, "crohn": str, "colitis": str}
    """
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return {}

    headers = [str(h).strip() if h else "" for h in rows[0]]
    try:
        code_idx = headers.index("Code")
        desc_idx = headers.index("Beschreibung")
        crohn_idx = headers.index("Bewertung_Crohn")
        colitis_idx = headers.index("Bewertung_Colitis")
    except ValueError as e:
        print(f"[!] Spalte nicht gefunden in {path}: {e}")
        wb.close()
        return {}

    result = {}
    for row in rows[1:]:
        code = str(row[code_idx]).strip() if row[code_idx] else ""
        if not code:
            continue
        desc = str(row[desc_idx]).strip() if row[desc_idx] else ""
        crohn = str(row[crohn_idx]).strip().lower() if row[crohn_idx] else "akzeptabel"
        colitis = str(row[colitis_idx]).strip().lower() if row[colitis_idx] else "akzeptabel"
        result[code] = {"beschreibung": desc, "crohn": crohn, "colitis": colitis}

    wb.close()
    return result


def load_food_config_both(path: Path) -> dict[str, dict]:
    """Load food config with both Crohn and Colitis ratings in one pass.

    Returns dict: name_lower -> {"name": str, "crohn": str, "colitis": str}
    """
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return {}

    headers = [str(h).strip() if h else "" for h in rows[0]]
    try:
        name_idx = headers.index("Nahrungsmittel")
        crohn_idx = headers.index("Bewertung_Crohn")
        colitis_idx = headers.index("Bewertung_Colitis")
    except ValueError as e:
        print(f"[!] Spalte nicht gefunden in {path}: {e}")
        wb.close()
        return {}

    result = {}
    for row in rows[1:]:
        name = str(row[name_idx]).strip() if row[name_idx] else ""
        if not name:
            continue
        crohn = str(row[crohn_idx]).strip().lower() if row[crohn_idx] else "akzeptabel"
        colitis = str(row[colitis_idx]).strip().lower() if row[colitis_idx] else "akzeptabel"
        result[name.lower()] = {"name": name, "crohn": crohn, "colitis": colitis}

    wb.close()
    return result
