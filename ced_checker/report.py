"""Generate xlsx reports for a date range with CED compatibility ratings."""

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from ced_checker.api import fetch_meals
from ced_checker.analyzer import analyze_and_rank
from ced_checker.models import MealRating


# --- Styling constants ---
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

GRADE_FILLS = {
    "A": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),  # green
    "B": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # light green
    "C": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # yellow
    "D": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),  # orange
    "E": PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid"),  # red-orange
    "F": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # red
}
GRADE_FONTS = {
    "F": Font(bold=True, color="FFFFFF"),
    "E": Font(bold=True, color="FFFFFF"),
}

SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=11, color="2F5496")


def _style_header(ws, row: int, max_col: int):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER


def _style_cell(cell, grade: str | None = None):
    """Apply border and optional grade color to a cell."""
    cell.border = BORDER
    if grade and grade in GRADE_FILLS:
        cell.fill = GRADE_FILLS[grade]
        if grade in GRADE_FONTS:
            cell.font = GRADE_FONTS[grade]


def generate_report(
    date_from: date,
    date_to: date,
    mode: str,
    settings: dict,
    allergen_config: dict,
    food_config: dict,
    output_path: Path,
) -> Path:
    """Fetch meals for each day in the range and write results to xlsx.

    Returns the path of the generated file.
    """
    base_url = settings["base_url"]
    api_path = settings["api_path"]
    locations = settings["locations"]
    mode_label = "Morbus Crohn" if mode == "crohn" else "Colitis ulcerosa"

    wb = Workbook()

    # ---- Sheet 1: Detail view ----
    ws_detail = wb.active
    ws_detail.title = "Detailübersicht"

    detail_headers = [
        "Datum", "Mensa", "Rang", "Note", "Score",
        "Gericht", "Preis (EUR)", "kcal", "Protein (g)",
        "Fett (g)", "Ballaststoffe (g)", "Salz (g)",
        "Tags", "Positiv", "Warnungen",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h)
    _style_header(ws_detail, 1, len(detail_headers))

    # ---- Sheet 2: Summary / best per day ----
    ws_summary = wb.create_sheet("Tagesübersicht")
    summary_headers = [
        "Datum", "Wochentag", "Beste Mensa", "Bestes Gericht",
        "Note", "Score", "Preis (EUR)", "Warnungen",
        "Gerichte gesamt", "Davon verträglich (A-C)", "Davon ausgeschlossen (F)",
    ]
    for col, h in enumerate(summary_headers, 1):
        ws_summary.cell(row=1, column=col, value=h)
    _style_header(ws_summary, 1, len(summary_headers))

    # ---- Sheet 3: Statistics ----
    ws_stats = wb.create_sheet("Statistik")

    detail_row = 2
    summary_row = 2

    day_count = (date_to - date_from).days + 1
    all_days_data = []  # for statistics

    WEEKDAYS_DE = ["Montag", "Dienstag", "Mittwoch", "Donnerstag",
                   "Freitag", "Samstag", "Sonntag"]

    current = date_from
    while current <= date_to:
        date_str = current.isoformat()
        weekday = WEEKDAYS_DE[current.weekday()]
        day_ratings = []  # all ratings for this day

        for loc in locations:
            meals = fetch_meals(base_url, api_path, loc["name"], date_str, loc["categories"])
            if not meals:
                continue

            ratings = analyze_and_rank(meals, allergen_config, food_config)

            for rank, r in enumerate(ratings, 1):
                grade_label = MealRating.GRADE_LABELS.get(r.grade, "")
                tags_str = ", ".join(r.meal.legend_tags)
                pos_str = ", ".join(r.positives)
                warn_str = "; ".join(r.warnings)

                values = [
                    date_str, loc["label"], rank,
                    f"{r.grade} ({grade_label})", r.score,
                    r.meal.title, r.meal.price, r.meal.calories,
                    r.meal.protein, r.meal.fat, r.meal.fiber, r.meal.salt,
                    tags_str, pos_str, warn_str,
                ]
                for col, val in enumerate(values, 1):
                    cell = ws_detail.cell(row=detail_row, column=col, value=val)
                    _style_cell(cell, r.grade if col in (3, 4, 5) else None)

                day_ratings.append((loc["label"], r))
                detail_row += 1

        # Day summary row
        all_days_data.append((current, day_ratings))

        if day_ratings:
            # best non-excluded meal of the day
            non_excluded = [(loc, r) for loc, r in day_ratings if not r.excluded]
            non_excluded.sort(key=lambda x: -x[1].score)

            total = len(day_ratings)
            good = sum(1 for _, r in day_ratings if r.grade in ("A", "B", "C"))
            excluded = sum(1 for _, r in day_ratings if r.grade == "F")

            if non_excluded:
                best_loc, best = non_excluded[0]
                best_grade_label = MealRating.GRADE_LABELS.get(best.grade, "")
                summary_values = [
                    date_str, weekday, best_loc, best.meal.title,
                    f"{best.grade} ({best_grade_label})", best.score,
                    best.meal.price, "; ".join(best.warnings) if best.warnings else "",
                    total, good, excluded,
                ]
            else:
                summary_values = [
                    date_str, weekday, "-", "Kein verträgliches Gericht",
                    "F", 0, 0, "", total, good, excluded,
                ]
        else:
            summary_values = [
                date_str, weekday, "-", "Keine Gerichte verfügbar",
                "-", 0, 0, "", 0, 0, 0,
            ]

        for col, val in enumerate(summary_values, 1):
            cell = ws_summary.cell(row=summary_row, column=col, value=val)
            cell.border = BORDER
            # Color the grade column
            if col == 5 and isinstance(val, str) and len(val) >= 1:
                grade_char = val[0]
                if grade_char in GRADE_FILLS:
                    cell.fill = GRADE_FILLS[grade_char]
                    if grade_char in GRADE_FONTS:
                        cell.font = GRADE_FONTS[grade_char]

        summary_row += 1
        current += timedelta(days=1)

    # ---- Build statistics sheet ----
    _build_statistics(ws_stats, all_days_data, mode_label, date_from, date_to)

    # ---- Auto-fit column widths ----
    for ws in [ws_detail, ws_summary, ws_stats]:
        _auto_width(ws)

    # ---- Freeze panes ----
    ws_detail.freeze_panes = "A2"
    ws_summary.freeze_panes = "A2"

    # ---- Auto-filter ----
    ws_detail.auto_filter.ref = f"A1:{get_column_letter(len(detail_headers))}1"
    ws_summary.auto_filter.ref = f"A1:{get_column_letter(len(summary_headers))}1"

    wb.save(output_path)
    return output_path


def _build_statistics(ws, all_days_data, mode_label, date_from, date_to):
    """Populate the statistics sheet with aggregate data."""
    row = 1

    # Title
    cell = ws.cell(row=row, column=1, value=f"CED Mensa-Checker Statistik — {mode_label}")
    cell.font = Font(bold=True, size=14, color="2F5496")
    row += 1
    ws.cell(row=row, column=1, value=f"Zeitraum: {date_from.isoformat()} bis {date_to.isoformat()}")
    row += 2

    # Collect all ratings
    all_ratings = []
    for day_date, day_ratings in all_days_data:
        for loc, r in day_ratings:
            all_ratings.append((day_date, loc, r))

    if not all_ratings:
        ws.cell(row=row, column=1, value="Keine Daten im gewählten Zeitraum.")
        return

    total = len(all_ratings)
    grade_counts = {}
    for _, _, r in all_ratings:
        grade_counts[r.grade] = grade_counts.get(r.grade, 0) + 1

    # Grade distribution
    ws.cell(row=row, column=1, value="Notenverteilung")
    ws.cell(row=row, column=1).font = SUBHEADER_FONT
    row += 1
    stat_headers = ["Note", "Bedeutung", "Anzahl", "Anteil"]
    for col, h in enumerate(stat_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
    row += 1

    for grade in ["A", "B", "C", "D", "E", "F"]:
        count = grade_counts.get(grade, 0)
        pct = count / total if total > 0 else 0
        label = MealRating.GRADE_LABELS.get(grade, "")
        values = [grade, label, count, f"{pct:.1%}"]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            _style_cell(cell, grade if col == 1 else None)
        row += 1

    # Total
    ws.cell(row=row, column=1, value="Gesamt").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total).font = Font(bold=True)
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = BORDER
    row += 2

    # Most common warnings
    ws.cell(row=row, column=1, value="Häufigste Warnungen")
    ws.cell(row=row, column=1).font = SUBHEADER_FONT
    row += 1

    warning_counts = {}
    for _, _, r in all_ratings:
        for w in r.warnings:
            warning_counts[w] = warning_counts.get(w, 0) + 1

    warn_headers = ["Warnung", "Häufigkeit"]
    for col, h in enumerate(warn_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
    row += 1

    sorted_warnings = sorted(warning_counts.items(), key=lambda x: -x[1])[:15]
    for w_text, w_count in sorted_warnings:
        ws.cell(row=row, column=1, value=w_text).border = BORDER
        ws.cell(row=row, column=2, value=w_count).border = BORDER
        row += 1

    row += 1

    # Best meals overall
    ws.cell(row=row, column=1, value="Top 10 Gerichte im Zeitraum")
    ws.cell(row=row, column=1).font = SUBHEADER_FONT
    row += 1

    top_headers = ["Rang", "Datum", "Mensa", "Gericht", "Note", "Score"]
    for col, h in enumerate(top_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
    row += 1

    non_excluded = [(d, loc, r) for d, loc, r in all_ratings if not r.excluded]
    non_excluded.sort(key=lambda x: (-x[2].score, x[2].meal.title))
    for i, (d, loc, r) in enumerate(non_excluded[:10], 1):
        values = [i, d.isoformat(), loc, r.meal.title, r.grade, r.score]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            _style_cell(cell, r.grade if col == 5 else None)
        row += 1

    row += 1

    # Per-location stats
    ws.cell(row=row, column=1, value="Statistik nach Mensa")
    ws.cell(row=row, column=1).font = SUBHEADER_FONT
    row += 1

    loc_headers = ["Mensa", "Gerichte", "Ø Score", "Beste Note", "A+B Anteil"]
    for col, h in enumerate(loc_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
    row += 1

    loc_groups = {}
    for d, loc, r in all_ratings:
        if loc not in loc_groups:
            loc_groups[loc] = []
        loc_groups[loc].append(r)

    for loc_name, ratings in sorted(loc_groups.items()):
        count = len(ratings)
        avg_score = sum(r.score for r in ratings) / count if count else 0
        best_grade = min((r.grade for r in ratings), key=lambda g: "ABCDEF".index(g))
        ab_count = sum(1 for r in ratings if r.grade in ("A", "B"))
        ab_pct = ab_count / count if count else 0

        values = [loc_name, count, round(avg_score, 1), best_grade, f"{ab_pct:.1%}"]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = BORDER
        row += 1


def _auto_width(ws, min_width: int = 8, max_width: int = 50):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                cell_len = len(str(cell.value))
                max_len = max(max_len, min(cell_len + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len
